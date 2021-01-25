from openpyxl import Workbook, load_workbook

wb = load_workbook("inventory_activity_report.xlsx")
ws = wb.active
temp = ws['D8'].value
temp = temp.replace("‚Äú", '"')
temp = temp.replace("‚Äù", '"')
print(temp)